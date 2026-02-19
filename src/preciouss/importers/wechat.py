"""WeChat Pay (微信支付) CSV/XLSX importer."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from preciouss.importers.base import CsvImporter, Transaction
from preciouss.importers.clearing import detect_merchant_clearing, resolve_payment_to_clearing

_ACCEPTED_STATUS_EXACT = frozenset(
    {
        "支付成功", "已转账", "已存入零钱", "已收钱", "朋友已收钱", "已全额退款",
        "对方已收钱",  # transfer accepted by other party
        "对方已退还",  # money returned by other party
        "充值成功",    # recharge micro-rewards (income)
        "已到账",      # payment arrived (income)
    }
)
_RE_REFUND_AMOUNT = re.compile(r"已退款[（(]￥([\d.]+)[）)]")
_RE_INCOME_TOTAL = re.compile(r"收入：\d+笔\s*([\d.]+)元")
_RE_EXPENSE_TOTAL = re.compile(r"支出：\d+笔\s*([\d.]+)元")


def _accept_status(status: str) -> bool:
    return status in _ACCEPTED_STATUS_EXACT or status.startswith("已退款")


class WechatImporter(CsvImporter):
    """Import transactions from WeChat Pay CSV or XLSX exports.

    WeChat Pay format (after metadata lines):
    交易时间, 交易类型, 交易对方, 商品, 收/支, 金额(元), 支付方式, 当前状态,
    交易单号, 商户单号, 备注

    Notes:
    - Amounts have ¥ prefix
    - 交易单号 and 商户单号 may have trailing tabs
    - 支付方式 contains bank card info like "招商银行(0913)"
    - Supports both .csv and .xlsx files (same column structure)
    """

    expected_headers = ["微信支付账单明细"]

    def __init__(self, account: str = "Assets:WeChat", currency: str = "CNY"):
        self._account = account
        self._currency = currency

    def account_name(self) -> str:
        return self._account

    def identify(self, filepath) -> bool:
        filepath = Path(filepath)
        suffix = filepath.suffix.lower()
        if suffix == ".csv":
            return self._identify_csv(filepath)
        elif suffix == ".xlsx":
            return self._identify_xlsx(filepath)
        return False

    def extract(self, filepath) -> list[Transaction]:
        filepath = Path(filepath)
        suffix = filepath.suffix.lower()
        if suffix == ".csv":
            return self._extract_csv(filepath)
        elif suffix == ".xlsx":
            return self._extract_xlsx(filepath)
        return []

    # --- CSV methods ---

    def _identify_csv(self, filepath: Path) -> bool:
        try:
            content = self._read_file(filepath)
            first_line = content.split("\n")[0]
            return "微信支付账单明细" in first_line
        except Exception:
            return False

    def _extract_csv(self, filepath: Path) -> list[Transaction]:
        """Extract transactions from CSV, finding header dynamically."""
        content = self._read_file(filepath)
        lines = content.split("\n")

        # Find the header line containing "交易时间"
        header_idx = None
        for i, line in enumerate(lines):
            if line.startswith("交易时间,"):
                header_idx = i
                break

        if header_idx is None:
            return []

        header_lines = lines[:header_idx]
        csv_lines = lines[header_idx:]
        csv_content = "\n".join(csv_lines)

        reader = csv.DictReader(io.StringIO(csv_content))
        transactions = []
        for row in reader:
            row = {k.strip(): v.strip() if v else "" for k, v in row.items() if k}
            tx = self._parse_row(row)
            if tx is not None:
                transactions.append(tx)

        totals = self._parse_header_totals(header_lines)
        if totals is not None:
            self._validate_totals(transactions, totals[0], totals[1], filepath)

        return transactions

    # --- XLSX methods ---

    def _identify_xlsx(self, filepath: Path) -> bool:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return False

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            # Check first row first cell for "微信支付"
            for row in ws.iter_rows(max_row=1, max_col=1, values_only=True):
                val = str(row[0]) if row[0] is not None else ""
                wb.close()
                return "微信支付" in val
            wb.close()
            return False
        except Exception:
            return False

    def _extract_xlsx(self, filepath: Path) -> list[Transaction]:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl is required for xlsx support. Install it with: uv add openpyxl"
            ) from e

        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        headers = None
        transactions = []
        header_buffer: list[str] = []

        for row in ws.iter_rows(values_only=True):
            # Convert all cells to strings
            cells = [str(c).strip() if c is not None else "" for c in row]

            # Find header row
            if headers is None:
                if cells and cells[0] == "交易时间":
                    headers = cells
                else:
                    header_buffer.append(" ".join(cells))
                continue

            # Build dict from row using headers
            if len(cells) < len(headers):
                continue
            row_dict = {headers[i]: cells[i] for i in range(len(headers))}
            tx = self._parse_row(row_dict)
            if tx is not None:
                transactions.append(tx)

        wb.close()

        totals = self._parse_header_totals(header_buffer)
        if totals is not None:
            self._validate_totals(transactions, totals[0], totals[1], filepath)

        return transactions

    # --- Header total validation ---

    @staticmethod
    def _parse_header_totals(lines: list[str]) -> tuple[Decimal, Decimal] | None:
        income_total = None
        expense_total = None
        for line in lines:
            m = _RE_INCOME_TOTAL.search(line)
            if m:
                income_total = Decimal(m.group(1))
            m = _RE_EXPENSE_TOTAL.search(line)
            if m:
                expense_total = Decimal(m.group(1))
        if income_total is not None and expense_total is not None:
            return income_total, expense_total
        return None

    def _validate_totals(
        self,
        transactions: list[Transaction],
        expected_income: Decimal,
        expected_expense: Decimal,
        filepath: Path,
    ) -> None:
        actual_income = sum(tx.amount for tx in transactions if tx.tx_type == "income")
        actual_expense = sum(-tx.amount for tx in transactions if tx.tx_type == "expense")
        if abs(actual_income - expected_income) > Decimal("0.01"):
            raise ValueError(
                f"[{filepath.name}] income {actual_income} ≠ expected {expected_income}"
            )
        if abs(actual_expense - expected_expense) > Decimal("0.01"):
            raise ValueError(
                f"[{filepath.name}] expense {actual_expense} ≠ expected {expected_expense}"
            )

    # --- Shared parsing ---

    def _parse_row(self, row: dict[str, str]) -> Transaction | None:
        # Parse date
        date_str = row.get("交易时间", "").strip()
        if not date_str:
            return None

        try:
            date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None

        # Parse status - skip non-completed transactions
        status = row.get("当前状态", "").strip()
        if not _accept_status(status):
            return None

        # Extract partial refund amount from status if present
        refund_amount: str | None = None
        m = _RE_REFUND_AMOUNT.match(status)
        if m:
            refund_amount = m.group(1)

        # Parse amount - remove ¥ prefix
        amount_str = row.get("金额(元)", "").strip()
        amount_str = amount_str.replace("¥", "").replace(",", "").strip()
        try:
            amount = Decimal(amount_str)
        except (InvalidOperation, ValueError):
            return None

        # Direction
        direction = row.get("收/支", "").strip()
        if direction == "支出":
            amount = -abs(amount)
            tx_type = "expense"
        elif direction == "收入":
            amount = abs(amount)
            tx_type = "income"
        elif direction == "/":
            tx_type = "transfer"
        else:
            tx_type = "other"

        payee = row.get("交易对方", "").strip()
        narration = row.get("商品", "").strip().strip('"')
        payment_method = row.get("支付方式", "").strip()
        trade_no = row.get("交易单号", "").strip().strip("\t")
        merchant_no = row.get("商户单号", "").strip().strip("\t")
        tx_type_raw = row.get("交易类型", "").strip()

        # Resolve payment account via clearing
        if payment_method and payment_method not in ("", "/"):
            resolved_account = resolve_payment_to_clearing(payment_method, "WX")
        else:
            resolved_account = self._account

        # Detect known merchants → counter_account (clearing)
        counter_account = detect_merchant_clearing("WX", payee, narration)

        return Transaction(
            date=date,
            amount=amount,
            currency=self._currency,
            payee=payee,
            narration=narration,
            source_account=resolved_account,
            payment_method=payment_method if payment_method and payment_method != "/" else None,
            reference_id=trade_no if trade_no and trade_no != "/" else None,
            counterpart_ref=merchant_no if merchant_no and merchant_no != "/" else None,
            raw_category=tx_type_raw or None,
            tx_type=tx_type,
            counter_account=counter_account,
            metadata={
                "wechat_status": status,
                "wechat_type": tx_type_raw,
                **({"wechat_refund_amount": refund_amount} if refund_amount else {}),
            },
        )
